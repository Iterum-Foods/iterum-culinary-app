#!/usr/bin/env python3
"""
Missing Information Detector
Identifies what information needs to be added to recipes for complete Iterum format
"""

import sqlite3
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import json

class MissingInfoDetector:
    """Detects missing information in recipes for Iterum format."""
    
    # Required fields for complete Iterum recipe
    REQUIRED_FIELDS = {
        'header': {
            'recipe_name': 'Recipe name',
            'concept': 'Concept',
            'submitted_by': 'Submitted by',
            'number_of_portions': 'Number of Portions',
            'cuisine': 'Cuisine',
            'category': 'Category',
            'date': 'Date',
            'serving_size': 'Serving Size Per Person'
        },
        'ingredients': {
            'ingredient_name': 'Ingredient name',
            'quantity': 'Quantity',
            'weight': 'Weight',
            'volume': 'Volume',
            'ap_cost': 'AP$ / Unit',
            'unit': 'Unit',
            'yield_pct': 'Yield %',
            'ep_cost': 'EP$ / Unit',
            'total_cost': 'Cost'
        },
        'method': {
            'instructions': 'Method/Instructions'
        }
    }
    
    def __init__(self, library_path: str = "recipe_library", converted_path: str = "converted_iterum"):
        self.library_path = Path(library_path)
        self.converted_path = Path(converted_path)
        self.db_path = self.library_path / "recipe_library.db"
        self.converted_path.mkdir(exist_ok=True)
    
    def analyze_recipe(self, recipe_id: Optional[str] = None, file_path: Optional[Path] = None) -> Dict[str, Any]:
        """
        Analyze a recipe to identify missing information.
        
        Args:
            recipe_id: Recipe ID from database
            file_path: Path to recipe file (Excel, etc.)
            
        Returns:
            Dictionary with missing information analysis
        """
        if recipe_id:
            # Get recipe from database
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM recipes WHERE id = ?", (recipe_id,))
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                return {'error': 'Recipe not found in database'}
            
            # Get library file path
            library_file = Path(row[16])  # library_path column
            if not library_file.exists():
                return {'error': 'Recipe file not found in library'}
            
            file_path = library_file
        
        if not file_path or not file_path.exists():
            return {'error': 'File path not provided or does not exist'}
        
        # Analyze the file
        missing_info = {
            'file_path': str(file_path),
            'recipe_name': file_path.stem,
            'missing_fields': [],
            'incomplete_sections': [],
            'warnings': [],
            'completeness_score': 0.0
        }
        
        # Try to read as Excel
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # Check header fields
                header_issues = self._check_header_fields(ws)
                missing_info['missing_fields'].extend(header_issues)
                
                # Check ingredients
                ingredient_issues = self._check_ingredients(ws)
                missing_info['missing_fields'].extend(ingredient_issues)
                
                # Check method
                method_issues = self._check_method(ws)
                missing_info['missing_fields'].extend(method_issues)
                
                # Calculate completeness
                total_required = sum(len(fields) for fields in self.REQUIRED_FIELDS.values())
                missing_count = len(missing_info['missing_fields'])
                missing_info['completeness_score'] = max(0, (total_required - missing_count) / total_required * 100)
                
            except Exception as e:
                missing_info['error'] = f"Error reading Excel file: {str(e)}"
        
        # Check if converted file exists
        converted_file = self.converted_path / f"{file_path.stem}.xlsx"
        if not converted_file.exists():
            missing_info['warnings'].append("Recipe not yet converted to Iterum format")
        
        return missing_info
    
    def _check_header_fields(self, ws) -> List[Dict[str, Any]]:
        """Check header section for missing fields."""
        issues = []
        
        # Check for recipe name (usually in B3)
        if not ws['B3'].value or str(ws['B3'].value).strip() in ['', 'Untitled Recipe']:
            issues.append({
                'section': 'header',
                'field': 'recipe_name',
                'label': 'Recipe name',
                'location': 'B3',
                'severity': 'high'
            })
        
        # Check for concept (B4)
        if not ws['B4'].value or str(ws['B4'].value).strip() == '':
            issues.append({
                'section': 'header',
                'field': 'concept',
                'label': 'Concept',
                'location': 'B4',
                'severity': 'medium'
            })
        
        # Check for cuisine (H4)
        if not ws['H4'].value or str(ws['H4'].value).strip().lower() in ['unknown', '']:
            issues.append({
                'section': 'header',
                'field': 'cuisine',
                'label': 'Cuisine',
                'location': 'H4',
                'severity': 'medium'
            })
        
        # Check for number of portions (B6)
        if not ws['B6'].value or ws['B6'].value == 0:
            issues.append({
                'section': 'header',
                'field': 'number_of_portions',
                'label': 'Number of Portions',
                'location': 'B6',
                'severity': 'high'
            })
        
        return issues
    
    def _check_ingredients(self, ws) -> List[Dict[str, Any]]:
        """Check ingredients section for missing information."""
        issues = []
        
        # Find ingredients table (usually starts around row 14)
        ingredients_found = False
        ingredient_count = 0
        
        for row in range(14, min(100, ws.max_row + 1)):
            ingredient_name = ws[f'A{row}'].value
            
            if ingredient_name and str(ingredient_name).strip().lower() not in ['', 'ingredients', 'method', 'instructions']:
                ingredients_found = True
                ingredient_count += 1
                
                # Check if ingredient has required fields
                weight = ws[f'C{row}'].value
                volume = ws[f'D{row}'].value
                ap_cost = ws[f'E{row}'].value
                unit = ws[f'F{row}'].value
                yield_pct = ws[f'G{row}'].value
                ep_cost = ws[f'H{row}'].value
                total_cost = ws[f'I{row}'].value
                
                # Check for missing costing information
                if not ap_cost or str(ap_cost).strip() == '':
                    issues.append({
                        'section': 'ingredients',
                        'field': 'ap_cost',
                        'label': f'AP$ / Unit for {ingredient_name}',
                        'location': f'E{row}',
                        'severity': 'high',
                        'ingredient': str(ingredient_name)
                    })
                
                if not unit or str(unit).strip() == '':
                    issues.append({
                        'section': 'ingredients',
                        'field': 'unit',
                        'label': f'Unit for {ingredient_name}',
                        'location': f'F{row}',
                        'severity': 'high',
                        'ingredient': str(ingredient_name)
                    })
                
                if not yield_pct or str(yield_pct).strip() == '':
                    issues.append({
                        'section': 'ingredients',
                        'field': 'yield_pct',
                        'label': f'Yield % for {ingredient_name}',
                        'location': f'G{row}',
                        'severity': 'medium',
                        'ingredient': str(ingredient_name)
                    })
        
        if not ingredients_found:
            issues.append({
                'section': 'ingredients',
                'field': 'ingredients_table',
                'label': 'Ingredients table',
                'location': 'Row 14+',
                'severity': 'high'
            })
        
        return issues
    
    def _check_method(self, ws) -> List[Dict[str, Any]]:
        """Check method/instructions section."""
        issues = []
        
        # Find method section (usually after ingredients)
        method_found = False
        
        for row in range(14, min(200, ws.max_row + 1)):
            cell_value = ws[f'A{row}'].value
            if cell_value and 'method' in str(cell_value).lower():
                # Check if there are instructions after this
                has_instructions = False
                for next_row in range(row + 1, min(row + 20, ws.max_row + 1)):
                    if ws[f'A{next_row}'].value and str(ws[f'A{next_row}'].value).strip():
                        has_instructions = True
                        break
                
                if not has_instructions:
                    issues.append({
                        'section': 'method',
                        'field': 'instructions',
                        'label': 'Method/Instructions',
                        'location': f'A{row + 1}+',
                        'severity': 'high'
                    })
                
                method_found = True
                break
        
        if not method_found:
            issues.append({
                'section': 'method',
                'field': 'method_section',
                'label': 'Method section',
                'location': 'After ingredients',
                'severity': 'high'
            })
        
        return issues
    
    def analyze_all_recipes(self) -> Dict[str, Any]:
        """Analyze all recipes in the library for missing information."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, title, library_path FROM recipes")
        recipes = cursor.fetchall()
        conn.close()
        
        results = {
            'total_recipes': len(recipes),
            'analyzed_recipes': [],
            'summary': {
                'high_priority': 0,
                'medium_priority': 0,
                'low_priority': 0,
                'average_completeness': 0.0
            }
        }
        
        total_completeness = 0.0
        
        for recipe_id, title, library_path in recipes:
            analysis = self.analyze_recipe(recipe_id=recipe_id)
            
            if 'error' not in analysis:
                results['analyzed_recipes'].append(analysis)
                
                # Count by severity
                for issue in analysis['missing_fields']:
                    if issue['severity'] == 'high':
                        results['summary']['high_priority'] += 1
                    elif issue['severity'] == 'medium':
                        results['summary']['medium_priority'] += 1
                    else:
                        results['summary']['low_priority'] += 1
                
                total_completeness += analysis['completeness_score']
        
        if results['total_recipes'] > 0:
            results['summary']['average_completeness'] = total_completeness / results['total_recipes']
        
        return results
    
    def generate_missing_info_report(self, output_file: Optional[str] = None) -> str:
        """Generate a detailed report of missing information."""
        analysis = self.analyze_all_recipes()
        
        report_lines = [
            "=" * 80,
            "MISSING INFORMATION REPORT",
            "=" * 80,
            f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"\nTotal Recipes Analyzed: {analysis['total_recipes']}",
            f"Average Completeness: {analysis['summary']['average_completeness']:.1f}%",
            f"\nSummary:",
            f"  High Priority Issues: {analysis['summary']['high_priority']}",
            f"  Medium Priority Issues: {analysis['summary']['medium_priority']}",
            f"  Low Priority Issues: {analysis['summary']['low_priority']}",
            "\n" + "=" * 80,
            "\nDETAILED FINDINGS:\n"
        ]
        
        for recipe_analysis in analysis['analyzed_recipes']:
            report_lines.append(f"\nRecipe: {recipe_analysis['recipe_name']}")
            report_lines.append(f"Completeness: {recipe_analysis['completeness_score']:.1f}%")
            report_lines.append(f"File: {recipe_analysis['file_path']}")
            
            if recipe_analysis['missing_fields']:
                report_lines.append("\nMissing Information:")
                for issue in recipe_analysis['missing_fields']:
                    severity_icon = "🔴" if issue['severity'] == 'high' else "🟡" if issue['severity'] == 'medium' else "🟢"
                    report_lines.append(f"  {severity_icon} {issue['label']} ({issue['section']}) - Location: {issue['location']}")
            else:
                report_lines.append("  ✓ All required information present")
            
            if recipe_analysis['warnings']:
                report_lines.append("\nWarnings:")
                for warning in recipe_analysis['warnings']:
                    report_lines.append(f"  ⚠ {warning}")
            
            report_lines.append("-" * 80)
        
        report_text = "\n".join(report_lines)
        
        if output_file:
            output_path = Path(output_file)
            output_path.write_text(report_text, encoding='utf-8')
            return str(output_path)
        
        return report_text

def main():
    """Command line interface."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Missing Information Detector')
    parser.add_argument('--recipe-id', help='Analyze specific recipe by ID')
    parser.add_argument('--file', help='Analyze specific file')
    parser.add_argument('--all', action='store_true', help='Analyze all recipes')
    parser.add_argument('--report', help='Generate report file')
    
    args = parser.parse_args()
    
    detector = MissingInfoDetector()
    
    if args.recipe_id:
        result = detector.analyze_recipe(recipe_id=args.recipe_id)
        print(json.dumps(result, indent=2))
    
    elif args.file:
        result = detector.analyze_recipe(file_path=Path(args.file))
        print(json.dumps(result, indent=2))
    
    elif args.all:
        if args.report:
            report_path = detector.generate_missing_info_report(args.report)
            print(f"Report generated: {report_path}")
        else:
            report_text = detector.generate_missing_info_report()
            print(report_text)
    else:
        # Interactive mode
        print("\n" + "="*80)
        print("MISSING INFORMATION DETECTOR")
        print("="*80)
        print("\nAnalyzing all recipes...\n")
        
        report_text = detector.generate_missing_info_report()
        print(report_text)
        
        save = input("\nSave report to file? (y/n): ").strip().lower()
        if save == 'y':
            filename = f"missing_info_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            report_path = detector.generate_missing_info_report(filename)
            print(f"\nReport saved to: {report_path}")

if __name__ == "__main__":
    main()



